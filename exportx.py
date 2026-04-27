"""
pages/3_Export.py — Export search results to PDF (sales-ready list view).
"""
import textwrap
import streamlit as st
import sys, os, io
from datetime import datetime
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from config import DB_PATH
import sqlite3, json


# ── Load data ─────────────────────────────────────────────────────────────────
def load_for_export(priority_filter=None, min_score=0, kw_filter="", company_filter=""):
    try:
        conn = sqlite3.connect(DB_PATH, check_same_thread=False)
        conn.row_factory = sqlite3.Row
        q = "SELECT * FROM leads WHERE score >= ?"
        p = [min_score]
        if priority_filter:
            q += " AND priority IN (" + ",".join("?"*len(priority_filter)) + ")"
            p += priority_filter
        if kw_filter.strip():
            k = "%" + kw_filter.strip().lower() + "%"
            q += " AND (LOWER(title) LIKE ? OR LOWER(company) LIKE ? OR LOWER(search_kw) LIKE ?)"
            p += [k, k, k]
        if company_filter.strip():
            c = "%" + company_filter.strip().lower() + "%"
            q += " AND LOWER(company) LIKE ?"
            p += [c]
        q += " ORDER BY score DESC"
        rows = conn.execute(q, p).fetchall()
        conn.close()
        return [dict(r) for r in rows]
    except Exception:
        return []


def db_stats_export():
    try:
        conn = sqlite3.connect(DB_PATH, check_same_thread=False)
        total = conn.execute("SELECT COUNT(*) FROM leads").fetchone()[0]
        last  = conn.execute("SELECT MAX(saved_at) FROM leads").fetchone()[0]
        conn.close()
        return total, (last or "")[:10]
    except Exception:
        return 0, ""


# ── Helper: parse factors JSON ────────────────────────────────────────────────
def _parse_factors(job):
    """Returns dict of factor tuples: {key: (score, label)}"""
    raw = job.get("factors", {})
    if isinstance(raw, str):
        try:
            raw = json.loads(raw)
        except Exception:
            raw = {}
    return raw if isinstance(raw, dict) else {}


def _factor_summary(factors):
    """One-line summary of all factor scores, e.g. role=2/3 · intent=3/3 · ..."""
    MAXES = {"role_relevance": 3, "hiring_intent": 3,
             "company_fit": 2, "remote_signal": 2, "buying_trigger": 3}
    LABELS = {"role_relevance": "role", "hiring_intent": "intent",
              "company_fit": "fit", "remote_signal": "remote", "buying_trigger": "trigger"}
    parts = []
    for k, mx in MAXES.items():
        val_tuple = factors.get(k, (0, ""))
        val = val_tuple[0] if isinstance(val_tuple, (list, tuple)) else val_tuple
        lbl = LABELS.get(k, k)
        parts.append(f"{lbl}={val}/{mx}")
    return " · ".join(parts)


def _factor_detail(factors):
    """Multi-line breakdown with label text, e.g. 'Role relevance: 2/3 — strong role match'"""
    MAXES  = {"role_relevance": 3, "hiring_intent": 3,
               "company_fit": 2, "remote_signal": 2, "buying_trigger": 3}
    NAMES  = {"role_relevance": "Role relevance",
               "hiring_intent":  "Hiring intent",
               "company_fit":    "Company fit",
               "remote_signal":  "Remote signal",
               "buying_trigger": "Buying trigger"}
    lines = []
    for k, mx in MAXES.items():
        val_tuple = factors.get(k, (0, ""))
        if isinstance(val_tuple, (list, tuple)) and len(val_tuple) >= 2:
            val, txt = val_tuple[0], val_tuple[1]
        else:
            val, txt = (val_tuple if isinstance(val_tuple, int) else 0), ""
        line = f"{NAMES.get(k, k)}: {val}/{mx}"
        if txt:
            line += f" — {txt}"
        lines.append(line)
    return "\n".join(lines)


# ── PDF builder ───────────────────────────────────────────────────────────────
def build_pdf(jobs, title="Lead Report", include_reason=True, include_trace=False):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                        Table, TableStyle, HRFlowable)
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    except ImportError:
        return None, "reportlab not installed"

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=18*mm, rightMargin=18*mm,
        topMargin=18*mm, bottomMargin=18*mm,
    )
    styles = getSampleStyleSheet()

    H1     = ParagraphStyle("H1",  fontSize=18, fontName="Helvetica-Bold",
                             textColor=colors.HexColor("#111827"), spaceAfter=4)
    SUB    = ParagraphStyle("SUB", fontSize=9,  fontName="Helvetica",
                             textColor=colors.HexColor("#6b7280"), spaceAfter=16)
    TH     = ParagraphStyle("TH",  fontSize=7.5, fontName="Helvetica-Bold",
                             textColor=colors.HexColor("#374151"))
    TD     = ParagraphStyle("TD",  fontSize=8,  fontName="Helvetica",
                             textColor=colors.HexColor("#111827"), leading=11)
    TDS    = ParagraphStyle("TDS", fontSize=7.5, fontName="Helvetica",
                             textColor=colors.HexColor("#6b7280"), leading=10)
    REASON = ParagraphStyle("RSN", fontSize=7.5, fontName="Helvetica-Oblique",
                             textColor=colors.HexColor("#374151"),
                             backColor=colors.HexColor("#fffbeb"),
                             borderPadding=(3,5,3,5), leading=10)

    PRIORITY_COLORS = {
        "High":   colors.HexColor("#fef2f2"),
        "Medium": colors.HexColor("#fffbeb"),
        "Low":    colors.HexColor("#f8fafc"),
    }
    PRIORITY_TEXT = {
        "High":   colors.HexColor("#b91c1c"),
        "Medium": colors.HexColor("#b45309"),
        "Low":    colors.HexColor("#64748b"),
    }
    SCORE_COLORS = {
        "High":   colors.HexColor("#ef4444"),
        "Medium": colors.HexColor("#f59e0b"),
        "Low":    colors.HexColor("#94a3b8"),
    }

    story = []
    now = datetime.now().strftime("%d %b %Y, %H:%M")

    story.append(Paragraph(title, H1))
    story.append(Paragraph(
        f"Generated: {now} &nbsp;·&nbsp; {len(jobs)} leads &nbsp;·&nbsp; "
        f"High: {sum(1 for j in jobs if j.get('priority')=='High')} &nbsp;·&nbsp; "
        f"Medium: {sum(1 for j in jobs if j.get('priority')=='Medium')}",
        SUB))
    story.append(HRFlowable(width="100%", thickness=0.5,
                             color=colors.HexColor("#e5e7eb"), spaceAfter=12))

    col_widths = [90, 110, 52, 65, 35]
    if include_reason:
        col_widths.append(175)

    header_row = [
        Paragraph("Company", TH),
        Paragraph("Role", TH),
        Paragraph("Location", TH),
        Paragraph("Priority", TH),
        Paragraph("Score", TH),
    ]
    if include_reason:
        header_row.append(Paragraph("Why they will buy", TH))

    table_data = [header_row]
    row_styles = []

    for i, job in enumerate(jobs):
        pri   = job.get("priority", "Low")
        score = job.get("score", 0)
        loc   = (job.get("location","") or "")[:24]
        url   = job.get("url","")

        title_cell = Paragraph(
            f'<a href="{url}" color="#111827"><b>{job.get("title","")}</b></a>'
            if url else f'<b>{job.get("title","")}</b>', TD)
        company_cell = Paragraph(job.get("company",""), TD)
        loc_cell     = Paragraph(loc, TDS)

        pri_style = ParagraphStyle("PRI", fontSize=7, fontName="Helvetica-Bold",
                                    textColor=PRIORITY_TEXT.get(pri, colors.black),
                                    alignment=TA_CENTER)
        pri_cell = Paragraph(pri.upper(), pri_style)

        sc_style = ParagraphStyle("SC", fontSize=11, fontName="Helvetica-Bold",
                                   textColor=SCORE_COLORS.get(pri, colors.gray),
                                   alignment=TA_CENTER)
        sc_cell = Paragraph(f"{score}/13", sc_style)

        row = [company_cell, title_cell, loc_cell, pri_cell, sc_cell]

        if include_reason:
            raw_reason = job.get("buy_reason","")
            if "→" in raw_reason:
                parts  = [p.strip() for p in raw_reason.split("→")]
                labels = ["Signal", "Meaning", "Why VE"]
                lines  = []
                for ri, part in enumerate(parts[:3]):
                    lbl = labels[ri] if ri < 3 else ""
                    lines.append(f"<b>{lbl}:</b> {part}")
                reason_text = " → ".join(lines)
            else:
                reason_text = raw_reason or "—"
            row.append(Paragraph(reason_text, REASON))

        table_data.append(row)
        bg = colors.HexColor("#fafafa") if i % 2 == 0 else colors.white
        row_styles.append(("BACKGROUND", (0, i+1), (-1, i+1), bg))
        row_styles.append(("BACKGROUND", (3, i+1), (3, i+1),
                            PRIORITY_COLORS.get(pri, colors.white)))

    table = Table(table_data, colWidths=[w*mm for w in col_widths], repeatRows=1)
    base_style = [
        ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor("#f1f5f9")),
        ("TEXTCOLOR",   (0, 0), (-1, 0), colors.HexColor("#374151")),
        ("FONTSIZE",    (0, 0), (-1, -1), 8),
        ("GRID",        (0, 0), (-1, -1), 0.3, colors.HexColor("#e5e7eb")),
        ("LINEBELOW",   (0, 0), (-1, 0),  0.8, colors.HexColor("#cbd5e1")),
        ("VALIGN",      (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING",  (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING",(0, 0), (-1, -1), 5),
    ] + row_styles

    table.setStyle(TableStyle(base_style))
    story.append(table)
    story.append(Spacer(1, 8*mm))
    story.append(HRFlowable(width="100%", thickness=0.3,
                             color=colors.HexColor("#e5e7eb"), spaceBefore=4))
    story.append(Paragraph(
        f"Lead Identification Framework V3 · VE · Exported {now}",
        ParagraphStyle("FTR", fontSize=7, fontName="Helvetica",
                       textColor=colors.HexColor("#9ca3af"), alignment=TA_CENTER)
    ))

    doc.build(story)
    buf.seek(0)
    return buf, None


# ── Excel builder ─────────────────────────────────────────────────────────────
def build_xlsx(jobs, title="Lead Report"):
    """
    3-sheet Excel:
      Sheet 1 — Leads         : full data, colour-coded, Signal/Meaning/Why split
      Sheet 2 — Signal Details : per-factor scores + label text for every lead
      Sheet 3 — Summary        : stats, priority breakdown, source breakdown
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None, "openpyxl not installed"

    wb  = Workbook()
    now = datetime.now().strftime("%d %b %Y, %H:%M")

    # ── Palette ───────────────────────────────────────────────────────────────
    CLR = {
        "h_dark":    "0F172A",
        "h_slate":   "1E293B",
        "h_fg":      "FFFFFF",
        "high_bg":   "FEF2F2", "high_fg":   "B91C1C",
        "med_bg":    "FFFBEB", "med_fg":    "B45309",
        "low_bg":    "F8FAFC", "low_fg":    "64748B",
        "alt":       "F9FAFB",
        "signal_bg": "EFF6FF", "signal_fg": "1E40AF",
        "meaning_bg":"F0FDF4", "meaning_fg":"065F46",
        "why_bg":    "FFFBEB", "why_fg":    "854D0E",
        "border":    "E2E8F0",
        "score_hi":  "EF4444", "score_med": "F59E0B", "score_low": "94A3B8",
        # factor colour tiers
        "f3_bg": "D1FAE5", "f3_fg": "065F46",   # score = max
        "f2_bg": "FEF9C3", "f2_fg": "854D0E",   # score = mid
        "f1_bg": "FEE2E2", "f1_fg": "991B1B",   # score = 1
        "f0_bg": "F1F5F9", "f0_fg": "94A3B8",   # score = 0
    }

    def fill(h):
        return PatternFill("solid", fgColor=h)

    def border():
        s = Side(style="thin", color=CLR["border"])
        return Border(left=s, right=s, top=s, bottom=s)

    def font(bold=False, color="111827", size=9, italic=False, underline=False):
        return Font(name="Arial", bold=bold, color=color, size=size,
                    italic=italic, underline="single" if underline else None)

    def hdr_cell(ws, row, col, value, bg=None, fg=None, size=9,
                 bold=True, align="center", wrap=False):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = font(bold=bold, color=(fg or CLR["h_fg"]), size=size)
        c.fill      = fill(bg or CLR["h_slate"])
        c.alignment = Alignment(horizontal=align, vertical="center",
                                wrap_text=wrap, indent=0)
        c.border    = border()
        return c

    def data_cell(ws, row, col, value, bg="FFFFFF", fg="111827", size=9,
                  bold=False, italic=False, align="left", wrap=False,
                  hyperlink=None, indent=1):
        c = ws.cell(row=row, column=col, value=value)
        
        c.font      = font(bold=bold, color=fg, size=size, italic=italic,
                           underline=bool(hyperlink))
        c.fill      = fill(bg)
        c.alignment = Alignment(horizontal=align, vertical="top",
                                wrap_text=wrap, indent=indent)
        c.border    = border()
        if hyperlink:
            c.hyperlink = hyperlink
        return c

    PRI_STYLES = {
        "High":   (CLR["high_bg"],  CLR["high_fg"]),
        "Medium": (CLR["med_bg"],   CLR["med_fg"]),
        "Low":    (CLR["low_bg"],   CLR["low_fg"]),
    }
    SCORE_FG = lambda s: CLR["score_hi"] if s >= 10 else (CLR["score_med"] if s >= 7 else CLR["score_low"])
    FACTOR_KEYS = ["role_relevance", "hiring_intent", "company_fit",
                   "remote_signal", "buying_trigger"]
    FACTOR_NAMES = ["Role relevance", "Hiring intent", "Company fit",
                    "Remote signal", "Buying trigger"]
    FACTOR_MAX   = [3, 3, 2, 2, 3]

    def factor_bg(val, mx):
        if val >= mx:          return CLR["f3_bg"], CLR["f3_fg"]
        elif val >= mx * 0.66: return CLR["f2_bg"], CLR["f2_fg"]
        elif val >= 1:         return CLR["f1_bg"], CLR["f1_fg"]
        else:                  return CLR["f0_bg"], CLR["f0_fg"]

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1 — Leads
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Leads"
    ws1.sheet_view.showGridLines = False
    ws1.freeze_panes = "A3"

    # Column layout
    # A  Company       B  Role         C  Location     D  Source
    # E  Posted        F  Priority     G  Score /13    H  Signal
    # I  Meaning       J  Why VE       K  Full Reason  L  Score Trace  M  URL
    COL_W = [20, 26, 16, 10, 10, 9, 9, 30, 30, 32, 50, 38, 22]
    HEADERS = [
        "Company", "Role", "Location", "Source", "Posted",
        "Priority", "Score /13",
        "Signal", "Meaning", "Why VE",
        "Full buy reason", "Score trace", "Job URL",
    ]
    NCOLS = len(HEADERS)

    # Row 1 — report banner
    ws1.merge_cells(f"A1:{get_column_letter(NCOLS)}1")
    c = ws1.cell(row=1, column=1,
                 value=f"{title}   |   {now}   |   {len(jobs)} leads")
    c.font      = font(bold=True, color="FFFFFF", size=12)
    c.fill      = fill(CLR["h_dark"])
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws1.row_dimensions[1].height = 28

    # Row 2 — column headers
    HEADER_BG = {
        7:  CLR["h_dark"],          # Signal
        8:  CLR["signal_fg"],
        9:  CLR["meaning_fg"],
        10: CLR["why_fg"],
    }
    for ci, (h, w) in enumerate(zip(HEADERS, COL_W), 1):
        bg = HEADER_BG.get(ci - 1, CLR["h_slate"])
        hdr_cell(ws1, 2, ci, h, bg=bg)
        ws1.column_dimensions[get_column_letter(ci)].width = w
    ws1.row_dimensions[2].height = 20

    for ri, job in enumerate(jobs, 3):
        pri   = job.get("priority", "Low")
        score = job.get("score", 0)
        pbg, pfg = PRI_STYLES.get(pri, (CLR["low_bg"], CLR["low_fg"]))
        row_bg = CLR["alt"] if ri % 2 == 0 else "FFFFFF"

        # Parse reason
        raw_reason = job.get("buy_reason", "") or ""
        if "→" in raw_reason:
            rparts = [p.strip() for p in raw_reason.split("→")]
            sig_txt  = rparts[0] if len(rparts) > 0 else ""
            mng_txt  = rparts[1] if len(rparts) > 1 else ""
            why_txt  = rparts[2] if len(rparts) > 2 else ""
        else:
            sig_txt  = raw_reason
            mng_txt  = ""
            why_txt  = ""

        url = job.get("url", "") or ""

        row_vals = [
            job.get("company", ""),
            job.get("title", ""),
            (job.get("location", "") or "")[:30],
            job.get("source", ""),
            (job.get("posted_at", "") or "")[:10],
            pri,
            score,
            sig_txt,
            mng_txt,
            why_txt,
            raw_reason,
            job.get("step_trace", "") or "",
            url,
        ]

        for ci, val in enumerate(row_vals, 1):
            col_idx = ci - 1
            # defaults
            bg, fg, bold, italic, wrap, align, hyper, sz = (
                row_bg, "111827", False, False, False, "left", None, 9
            )

            if col_idx == 5:   # Priority
                bg, fg, bold, align = pbg, pfg, True, "center"
            elif col_idx == 6: # Score
                bg  = pbg
                fg  = SCORE_FG(score)
                bold, sz, align = True, 11, "center"
            elif col_idx == 7: # Signal
                bg, fg, italic, wrap = CLR["signal_bg"], CLR["signal_fg"], False, True
            elif col_idx == 8: # Meaning
                bg, fg, italic, wrap = CLR["meaning_bg"], CLR["meaning_fg"], True, True
            elif col_idx == 9: # Why VE
                bg, fg, italic, wrap = CLR["why_bg"], CLR["why_fg"], True, True
            elif col_idx == 10: # Full reason
                bg, fg, wrap, sz = "FFFDF0", "374151", True, 8
            elif col_idx == 11: # Trace
                bg, fg, sz = "F8FAFC", "6B7280", 8
            elif col_idx == 12: # URL
                fg, hyper, sz = "1D4ED8", (url or None), 8

            data_cell(ws1, ri, ci, val,
                      bg=bg, fg=fg, bold=bold, italic=italic,
                      wrap=wrap, align=align, hyperlink=hyper, size=sz,
                      indent=(0 if align == "center" else 1))

        # Row height: taller if reason columns have content
        has_reason = any([sig_txt, mng_txt, why_txt])
        ws1.row_dimensions[ri].height = 52 if has_reason else 18

    # Totals row
    tr = len(jobs) + 3
    ws1.merge_cells(f"A{tr}:F{tr}")
    c = ws1.cell(row=tr, column=1, value=f"Total: {len(jobs)} leads")
    c.font = font(bold=True, color="374151", size=9)
    c.fill = fill("F1F5F9")
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws1.row_dimensions[tr].height = 18

    sc = ws1.cell(row=tr, column=7, value=f"=AVERAGE(G3:G{len(jobs)+2})")
    sc.font = font(bold=True, color="374151", size=9)
    sc.fill = fill("F1F5F9")
    sc.number_format = "0.0"
    sc.alignment = Alignment(horizontal="center")
    sc.border = border()

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2 — Signal Details (per-factor breakdown)
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Signal Details")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = "A3"

    # Columns: Company | Role | Priority | Score | then one col per factor (score + label)
    S2_FIXED_HDRS = ["Company", "Role", "Priority", "Score /13",
                     "ICP signals matched", "Capacity signals"]
    S2_FIXED_W    = [22, 26, 9, 9, 30, 30]
    # Then 5 factor columns: "Role relevance (score)", "Role relevance (label)", ...
    FACTOR_COLS = []
    FACTOR_W    = []
    for fn in FACTOR_NAMES:
        FACTOR_COLS += [f"{fn} — score", f"{fn} — detail"]
        FACTOR_W    += [14, 34]

    ALL_S2_HDRS = S2_FIXED_HDRS + FACTOR_COLS
    ALL_S2_W    = S2_FIXED_W + FACTOR_W
    S2_NCOLS    = len(ALL_S2_HDRS)

    # Banner
    ws2.merge_cells(f"A1:{get_column_letter(S2_NCOLS)}1")
    c2 = ws2.cell(row=1, column=1,
                  value=f"{title} — Signal Details   |   {now}   |   {len(jobs)} leads")
    c2.font      = font(bold=True, color="FFFFFF", size=12)
    c2.fill      = fill(CLR["h_dark"])
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws2.row_dimensions[1].height = 28

    # Header row
    for ci, (h, w) in enumerate(zip(ALL_S2_HDRS, ALL_S2_W), 1):
        # Factor score cols get a teal bg; factor detail cols get a lighter bg
        is_score_col  = ("— score"  in h)
        is_detail_col = ("— detail" in h)
        bg = (CLR["meaning_fg"] if is_score_col else
              CLR["f3_bg"].replace("D1", "B6") if is_detail_col else CLR["h_slate"])
        hdr_cell(ws2, 2, ci, h, bg=bg)
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.row_dimensions[2].height = 36

    # ICP signal lists (for quick matching display)
    ICP_STARTUP   = ["startup","early-stage","seed","series a","series b","founded in 20",
                     "recently founded","pre-seed","newly funded","venture-backed"]
    ICP_SCALING   = ["scaling","rapidly growing","fast-growing","hypergrowth","expanding team",
                     "growing team","team expansion","building out","hiring across","just raised"]
    ICP_REMOTE    = ["remote","distributed","work from anywhere","fully remote","remote-first",
                     "remote friendly","hybrid","async","globally distributed","location agnostic"]
    ICP_OUTSOURCE = ["lean team","small team","tight budget","cost-effective","flexible",
                     "fast turnaround","contractor","freelancer","outsource","crm admin"]
    CAP_SIGNALS   = ["immediately","urgently","asap","multiple openings","several positions",
                     "rapidly","building out","extra capacity","growing workload","new role",
                     "first hire","building the team","new product launch"]

    for ri, job in enumerate(jobs, 3):
        pri   = job.get("priority", "Low")
        score = job.get("score", 0)
        pbg, pfg = PRI_STYLES.get(pri, (CLR["low_bg"], CLR["low_fg"]))
        row_bg = CLR["alt"] if ri % 2 == 0 else "FFFFFF"
        factors = _parse_factors(job)

        # Build ICP matched signals from description
        desc = ((job.get("description","") or "") + " " +
                (job.get("title","") or "") + " " +
                (job.get("company","") or "")).lower()

        icp_hits = []
        for sig in ICP_STARTUP + ICP_SCALING + ICP_REMOTE + ICP_OUTSOURCE:
            if sig in desc and sig not in icp_hits:
                icp_hits.append(sig)
        cap_hits = [s for s in CAP_SIGNALS if s in desc]

        # Fixed columns
        fixed_vals = [
            job.get("company", ""),
            job.get("title", ""),
            pri,
            score,
            ", ".join(icp_hits[:8]) if icp_hits else "none detected",
            ", ".join(cap_hits[:6]) if cap_hits else "none detected",
        ]
        for ci, val in enumerate(fixed_vals, 1):
            col_idx = ci - 1
            if col_idx == 2:   # Priority
                data_cell(ws2, ri, ci, val, bg=pbg, fg=pfg, bold=True, align="center")
            elif col_idx == 3: # Score
                data_cell(ws2, ri, ci, val, bg=pbg, fg=SCORE_FG(score), bold=True, size=11, align="center")
            elif col_idx in (4, 5):  # ICP / capacity signals
                data_cell(ws2, ri, ci, val, bg="F0F9FF", fg="0369A1", wrap=True, size=8)
            else:
                data_cell(ws2, ri, ci, val, bg=row_bg)

        # Factor columns (pairs: score | detail)
        for fi, (fk, fn, fmx) in enumerate(zip(FACTOR_KEYS, FACTOR_NAMES, FACTOR_MAX)):
            val_tuple = factors.get(fk, (0, ""))
            fval = val_tuple[0] if isinstance(val_tuple, (list, tuple)) else int(val_tuple)
            ftxt = val_tuple[1] if isinstance(val_tuple, (list, tuple)) and len(val_tuple) > 1 else ""
            fbg, ffg = factor_bg(fval, fmx)

            score_ci  = len(S2_FIXED_HDRS) + fi * 2 + 1
            detail_ci = score_ci + 1

            # Score cell: "{val}/{max}"
            c_sc = ws2.cell(row=ri, column=score_ci, value=f"{fval}/{fmx}")
            c_sc.font      = font(bold=True, color=ffg, size=10)
            c_sc.fill      = fill(fbg)
            c_sc.alignment = Alignment(horizontal="center", vertical="top")
            c_sc.border    = border()

            # Detail cell: label text
            c_dt = ws2.cell(row=ri, column=detail_ci, value=ftxt or "—")
            c_dt.font      = font(color=ffg, size=8, italic=True)
            c_dt.fill      = fill(fbg)
            c_dt.alignment = Alignment(horizontal="left", vertical="top",
                                       wrap_text=True, indent=1)
            c_dt.border    = border()

        ws2.row_dimensions[ri].height = 36

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 3 — Summary
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Summary")
    ws3.sheet_view.showGridLines = False
    for col_letter, w in zip(["A","B","C"], [26, 14, 20]):
        ws3.column_dimensions[col_letter].width = w

    def s3(row, col, value, bold=False, bg=None, fg="111827", size=10,
           align="left", wrap=False):
        c = ws3.cell(row=row, column=col, value=value)
        c.font = font(bold=bold, color=fg, size=size)
        if bg: c.fill = fill(bg)
        c.alignment = Alignment(horizontal=align, vertical="center",
                                wrap_text=wrap, indent=1)
        c.border = border()
        return c

    # Banner
    ws3.merge_cells("A1:C1")
    s3(1, 1, f"{title} — Summary", bold=True, bg=CLR["h_dark"], fg="FFFFFF",
       size=13, align="left")
    ws3.row_dimensions[1].height = 30
    s3(2, 1, f"Generated: {now}", fg="6B7280", size=8)
    ws3.row_dimensions[2].height = 16

    # Stats header
    s3(4, 1, "Metric",      bold=True, bg=CLR["h_slate"], fg="FFFFFF")
    s3(4, 2, "Count",       bold=True, bg=CLR["h_slate"], fg="FFFFFF", align="center")
    s3(4, 3, "% of Total",  bold=True, bg=CLR["h_slate"], fg="FFFFFF", align="center")
    ws3.row_dimensions[4].height = 18

    high_n  = sum(1 for j in jobs if j.get("priority")=="High")
    med_n   = sum(1 for j in jobs if j.get("priority")=="Medium")
    low_n   = sum(1 for j in jobs if j.get("priority")=="Low")
    avg_sc  = round(sum(j.get("score",0) for j in jobs) / max(len(jobs),1), 1)
    total_n = len(jobs)

    stat_rows = [
        ("Total leads",        total_n, "100%",
         "F1F5F9", "374151"),
        ("High priority",      high_n,  f"{round(high_n/max(total_n,1)*100)}%",
         CLR["high_bg"], CLR["high_fg"]),
        ("Medium priority",    med_n,   f"{round(med_n/max(total_n,1)*100)}%",
         CLR["med_bg"],  CLR["med_fg"]),
        ("Low priority",       low_n,   f"{round(low_n/max(total_n,1)*100)}%",
         CLR["low_bg"],  CLR["low_fg"]),
        ("Average score (/13)",avg_sc,  "—",
         "F0FDF4", "065F46"),
    ]
    for i, (label, count, pct, bg, fg) in enumerate(stat_rows, 5):
        s3(i, 1, label, bg=bg, fg=fg, size=9)
        s3(i, 2, count, bg=bg, fg=fg, size=9, align="center")
        s3(i, 3, pct,   bg=bg, fg=fg, size=9, align="center")
        ws3.row_dimensions[i].height = 18

    # Score bands
    s3(11, 1, "Score band",    bold=True, bg=CLR["h_slate"], fg="FFFFFF")
    s3(11, 2, "Count",         bold=True, bg=CLR["h_slate"], fg="FFFFFF", align="center")
    s3(11, 3, "Range",         bold=True, bg=CLR["h_slate"], fg="FFFFFF", align="center")
    ws3.row_dimensions[11].height = 18

    bands = [("10–13", 10, 13, CLR["high_bg"], CLR["high_fg"]),
             ("7–9",   7,  9,  CLR["med_bg"],  CLR["med_fg"]),
             ("0–6",   0,  6,  CLR["low_bg"],  CLR["low_fg"])]
    for i, (label, lo, hi, bg, fg) in enumerate(bands, 12):
        cnt = sum(1 for j in jobs if lo <= j.get("score",0) <= hi)
        s3(i, 1, f"Score {label}", bg=bg, fg=fg, size=9)
        s3(i, 2, cnt,              bg=bg, fg=fg, size=9, align="center")
        s3(i, 3, label,            bg=bg, fg=fg, size=9, align="center")
        ws3.row_dimensions[i].height = 18

    # Factor averages
    s3(16, 1, "Avg factor score", bold=True, bg=CLR["h_slate"], fg="FFFFFF")
    s3(16, 2, "Avg (/max)",       bold=True, bg=CLR["h_slate"], fg="FFFFFF", align="center")
    s3(16, 3, "Max possible",     bold=True, bg=CLR["h_slate"], fg="FFFFFF", align="center")
    ws3.row_dimensions[16].height = 18

    for i, (fk, fn, fmx) in enumerate(zip(FACTOR_KEYS, FACTOR_NAMES, FACTOR_MAX), 17):
        scores = []
        for job in jobs:
            fac = _parse_factors(job)
            vt = fac.get(fk, (0,""))
            scores.append(vt[0] if isinstance(vt,(list,tuple)) else int(vt))
        avg = round(sum(scores)/max(len(scores),1), 1)
        fbg, ffg = factor_bg(avg, fmx)
        s3(i, 1, fn,  bg=fbg, fg=ffg, size=9)
        s3(i, 2, avg, bg=fbg, fg=ffg, size=9, align="center")
        s3(i, 3, fmx, bg=fbg, fg=ffg, size=9, align="center")
        ws3.row_dimensions[i].height = 18

    # Source breakdown
    row_start = 17 + len(FACTOR_KEYS) + 1
    s3(row_start,     1, "Source",  bold=True, bg=CLR["h_slate"], fg="FFFFFF")
    s3(row_start,     2, "Count",   bold=True, bg=CLR["h_slate"], fg="FFFFFF", align="center")
    ws3.row_dimensions[row_start].height = 18
    sources = {}
    for j in jobs:
        src = j.get("source","unknown")
        sources[src] = sources.get(src, 0) + 1
    for i, (src, cnt) in enumerate(sorted(sources.items(), key=lambda x:-x[1]), row_start+1):
        s3(i, 1, src.title(), bg=CLR["alt"], fg="374151", size=9)
        s3(i, 2, cnt,         bg=CLR["alt"], fg="374151", size=9, align="center")
        ws3.row_dimensions[i].height = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, None


# ── Streamlit UI ──────────────────────────────────────────────────────────────
def run():
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=DM+Mono:wght@400;500&family=DM+Sans:wght@300;400;500;600&display=swap');
    html,body,[class*="css"]{ font-family:'DM Sans',sans-serif; }
    #MainMenu,footer,header{ visibility:hidden; }
    .block-container{ padding:2rem 2rem 4rem; max-width:900px; }
    .pg-title { font-family:'DM Mono',monospace; font-size:1.4rem; font-weight:500; color:#111827; margin:0; }
    .pg-sub   { font-size:0.8rem; color:#6b7280; margin-top:.2rem; }
    .preview-card { background:#fff; border:1px solid #e5e7eb; border-radius:10px;
                    padding:1rem 1.25rem; margin-bottom:.6rem; }
    .preview-title { font-size:.9rem; font-weight:600; color:#111827; margin:0 0 2px; }
    .preview-co    { font-size:.8rem; color:#374151; margin:0; }
    .preview-score { font-family:'DM Mono',monospace; font-size:1.1rem; font-weight:500; }
    .preview-reason{ font-size:.75rem; color:#374151; background:#fffbeb; border:1px solid #fde68a;
                    border-left:3px solid #f59e0b; padding:6px 9px; border-radius:0 5px 5px 0;
                    margin-top:6px; line-height:1.5; }
    .stButton>button { background:#111827!important; color:#fff!important; border:none!important;
        border-radius:8px!important; font-weight:500!important; }
    .stButton>button:hover{ background:#1f2937!important; }
    </style>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div style="padding:1.5rem 0 1rem;border-bottom:1px solid #e5e7eb;margin-bottom:1.5rem;">
    <p class="pg-title">📄 export</p>
    <p class="pg-sub">Download search results as PDF or Excel · Sales-ready format</p>
    </div>
    """, unsafe_allow_html=True)

    total_db, last_save = db_stats_export()

    if total_db == 0:
        st.markdown("""
        <div style="text-align:center;padding:3rem 2rem;color:#9ca3af;">
        <div style="font-size:2rem;margin-bottom:.5rem;">🗄️</div>
        <div style="font-size:.9rem;font-weight:500;color:#4b5563;">No leads in database</div>
        <div>Run a search first, then come back to export</div>
        </div>
        """, unsafe_allow_html=True)
        st.stop()

    st.markdown(f"""
    <div style="background:#f0fdf4;border:1px solid #bbf7d0;border-radius:8px;
                padding:10px 14px;font-size:.78rem;color:#065f46;margin-bottom:1.5rem;">
    💾 <strong>{total_db}</strong> leads available · Last saved: {last_save}
    </div>
    """, unsafe_allow_html=True)

    # ── Filters ───────────────────────────────────────────────────────────────
    st.markdown("### Filter leads to export")
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        exp_priority = st.multiselect("Priority", ["High","Medium","Low"],
                                      default=["High","Medium"])
    with col2:
        exp_min_score = st.slider("Min score (/13)", 0, 13, 8)
    with col3:
        exp_kw = st.text_input("Keyword filter", placeholder="role or keyword…")
    with col4:
        exp_company = st.text_input("Company name", placeholder="e.g. Acme…")

    # ── Options ───────────────────────────────────────────────────────────────
    st.markdown("### Export options")
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        report_title = st.text_input("Report title", value="Lead Report — VE")
    with c2:
        include_reason = st.checkbox("Include buy reason", value=True)
    with c3:
        include_trace  = st.checkbox("Include scoring trace", value=False)
    with c4:
        include_signals = st.checkbox("Signal Details sheet (Excel)", value=True)

    # ── Load & preview ────────────────────────────────────────────────────────
    jobs = load_for_export(
        priority_filter=exp_priority if exp_priority else None,
        min_score=exp_min_score,
        kw_filter=exp_kw,
        company_filter=exp_company,
    )

    st.markdown(f"### Preview ({len(jobs)} leads)")

    if not jobs:
        st.warning("No leads match your filters — try lowering the min score or adding more priority tiers.")
    else:
        high_n = sum(1 for j in jobs if j.get("priority")=="High")
        med_n  = sum(1 for j in jobs if j.get("priority")=="Medium")
        low_n  = sum(1 for j in jobs if j.get("priority")=="Low")
        avg_sc = round(sum(j.get("score",0) for j in jobs)/max(len(jobs),1),1)

        st.markdown(f"""
        <div style="display:flex;gap:.75rem;padding:.7rem 1.1rem;background:#f9fafb;
                    border:1px solid #e5e7eb;border-radius:8px;margin-bottom:1rem;flex-wrap:wrap;">
        <div style="text-align:center;flex:1;">
            <div style="font-family:'DM Mono',monospace;font-size:1.1rem;font-weight:500;">{len(jobs)}</div>
            <div style="font-size:.62rem;color:#9ca3af;text-transform:uppercase;">Total</div>
        </div>
        <div style="text-align:center;flex:1;">
            <div style="font-family:'DM Mono',monospace;font-size:1.1rem;color:#ef4444;font-weight:500;">{high_n}</div>
            <div style="font-size:.62rem;color:#9ca3af;text-transform:uppercase;">High</div>
        </div>
        <div style="text-align:center;flex:1;">
            <div style="font-family:'DM Mono',monospace;font-size:1.1rem;color:#f59e0b;font-weight:500;">{med_n}</div>
            <div style="font-size:.62rem;color:#9ca3af;text-transform:uppercase;">Medium</div>
        </div>
        <div style="text-align:center;flex:1;">
            <div style="font-family:'DM Mono',monospace;font-size:1.1rem;color:#94a3b8;font-weight:500;">{low_n}</div>
            <div style="font-size:.62rem;color:#9ca3af;text-transform:uppercase;">Low</div>
        </div>
        <div style="text-align:center;flex:1;">
            <div style="font-family:'DM Mono',monospace;font-size:1.1rem;color:#059669;font-weight:500;">{avg_sc}</div>
            <div style="font-size:.62rem;color:#9ca3af;text-transform:uppercase;">Avg score</div>
        </div>
        </div>
        """, unsafe_allow_html=True)

        for job in jobs[:5]:
            p   = (job.get("priority","Low")).lower()
            sc  = job.get("score", 0)
            sc_c = {"high":"#ef4444","medium":"#f59e0b","low":"#94a3b8"}.get(p,"#94a3b8")
            raw_reason = job.get("buy_reason","") or ""
            factors = _parse_factors(job)
            factor_line = _factor_summary(factors)

            if "→" in raw_reason:
                parts = [x.strip() for x in raw_reason.split("→")]
                labels = ["Signal","Meaning","Why VE"]
                r_parts = " <span style='color:#f59e0b;font-weight:700;margin:0 3px;'>→</span> ".join(
                    f"<b style='font-size:.67rem;text-transform:uppercase;color:#92400e;'>{labels[i]}</b> {pt}"
                    for i, pt in enumerate(parts[:3])
                )
                reason_html = f'<div class="preview-reason">{r_parts}</div>'
            elif raw_reason:
                reason_html = f'<div class="preview-reason">{raw_reason}</div>'
            else:
                reason_html = ""

            meta = " · ".join(filter(None,[
                job.get("company",""),
                job.get("location",""),
                (job.get("posted_at","") or "")[:10]
            ]))

            st.markdown(textwrap.dedent(f"""
            <div class="preview-card" style="border-left:3px solid {sc_c};">
            <div style="display:flex;align-items:flex-start;gap:1rem;">
                <div style="flex:1;">
                    <p class="preview-title">{job.get("title","")}</p>
                    <p class="preview-co" style="margin-bottom:4px;">{meta}</p>
                    <p style="font-size:.7rem;color:#6b7280;font-family:monospace;margin:0 0 4px;">{factor_line}</p>
                    <p style="font-size:.7rem;">{reason_html if include_reason else ""}</p>
                </div>
                <div style="text-align:center;flex-shrink:0;width:50px;">
                <div class="preview-score" style="color:{sc_c};">{sc}</div>
                <div style="font-size:.62rem;color:#9ca3af;">/13</div>
                </div>
            </div>
            </div>
            """), unsafe_allow_html=True)

        if len(jobs) > 5:
            st.caption(f"+ {len(jobs)-5} more leads in the export")

        # ── Export buttons ─────────────────────────────────────────────────────
        st.markdown("---")
        col_pdf, col_xl, col_spacer = st.columns([2, 2, 3])

        with col_pdf:
            if st.button("📥 Download PDF", use_container_width=True):
                with st.spinner("Building PDF…"):
                    try:
                        import reportlab
                        pdf_buf, err = build_pdf(
                            jobs, title=report_title,
                            include_reason=include_reason,
                            include_trace=include_trace,
                        )
                        if err:
                            st.error(f"PDF error: {err}")
                        else:
                            filename = f"leads_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
                            st.download_button(
                                label="⬇️ Save PDF", data=pdf_buf,
                                file_name=filename, mime="application/pdf",
                                use_container_width=True,
                            )
                            st.success(f"✅ PDF ready — {len(jobs)} leads")
                    except ImportError:
                        st.error("reportlab not installed. Run: pip install reportlab")

        with col_xl:
            if st.button("📊 Download Excel", use_container_width=True):
                with st.spinner("Building Excel…"):
                    try:
                        xl_buf, err = build_xlsx(jobs, title=report_title)
                        if err:
                            st.error(f"Excel error: {err}")
                        else:
                            filename = f"leads_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                            st.download_button(
                                label="⬇️ Save Excel", data=xl_buf,
                                file_name=filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True,
                            )
                            st.success(f"✅ Excel ready — {len(jobs)} leads · 3 sheets")
                    except ImportError:
                        st.error("openpyxl not installed. Run: pip install openpyxl")